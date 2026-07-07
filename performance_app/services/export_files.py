from __future__ import annotations

from pathlib import Path

from flask import current_app
from openpyxl import Workbook

from performance_app.db import get_db
from performance_app.repositories.audit import write_audit_log
from performance_app.services.calculation_runner import list_cycle_results
from performance_app.routes.pages import STATUS_LABELS
from performance_app.domain.constants import (
    GROUP_MANAGEMENT,
    GROUP_EMPLOYEE_P1_3,
    GROUP_EMPLOYEE_P4_10,
    MANAGEMENT_LABELS,
    EMPLOYEE_LABELS,
    FORCE_MANAGEMENT_DIMENSIONS_EMPLOYEES,
)

RESULT_HEADERS = ("工号", "姓名", "四级部门", "职级", "加权分", "排序", "部门负责人初评结果", "加权计算结果", "最终微调结果", "状态")


def create_cycle_export(cycle_id: int, export_type: str, operator_id: str, operator_name: str) -> dict:
    records = list_cycle_results(cycle_id)
    if not records:
        raise ValueError("no calculated results to export")

    export_id = f"cycle-{cycle_id}-{export_type}"
    export_dir = Path(current_app.config["EXPORT_DIR"]).resolve()
    export_dir.mkdir(parents=True, exist_ok=True)
    file_path = export_dir / f"{export_id}.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "结果总览"
    sheet.append(RESULT_HEADERS)
    for record in records:
        sheet.append(
            (
                record["emp_id"],
                record["emp_name"],
                record.get("dept_level_4") or record.get("dept_name") or "-",
                record.get("level") or "-",
                record.get("weighted_score") or "-",
                record.get("rank_in_group") or "-",
                record.get("current_subjective_level") or "-",
                record.get("suggested_level") or "-",
                record.get("final_level") or "-",
                STATUS_LABELS.get(record.get("status"), record.get("status") or "-"),
            )
        )
    workbook.save(file_path)

    write_audit_log(
        action="EXPORT_EXCEL",
        target_type="export_file",
        target_id=export_id,
        operator_id=operator_id,
        operator_name=operator_name,
        cycle_id=cycle_id,
        after_snapshot={"export_id": export_id, "export_type": export_type, "file_name": file_path.name},
    )
    get_db().commit()
    return {"export_id": export_id, "file_name": file_path.name, "download_url": f"/exports/{export_id}/download"}


# 基础字段（所有sheet共有）
BASE_HEADERS = ("工号", "姓名", "三级部门", "四级部门", "职级",
               "勤奋总量", "勤奋月均", "勤奋等级",
               "纪律异常次数", "纪律等级",
               "学习时长(h)", "学习排名%", "学习等级")

# 管理序列主观评价标签
MGMT_SUBJECTIVE_LABELS = (
    f"主观评价-{MANAGEMENT_LABELS['label_1']}",
    f"主观评价-{MANAGEMENT_LABELS['label_2']}",
    f"主观评价-{MANAGEMENT_LABELS['label_3']}",
)

# 员工序列主观评价标签
EMP_SUBJECTIVE_LABELS = (
    f"主观评价-{EMPLOYEE_LABELS['label_1']}",
    f"主观评价-{EMPLOYEE_LABELS['label_2']}",
    f"主观评价-{EMPLOYEE_LABELS['label_3']}",
)

COMMON_TAIL = ("部门负责人确认等级",)

# 管理序列表头
MGMT_HEADERS = BASE_HEADERS + MGMT_SUBJECTIVE_LABELS + COMMON_TAIL

# 员工序列表头
EMP_HEADERS = BASE_HEADERS + EMP_SUBJECTIVE_LABELS + COMMON_TAIL

PROCESS_QUERY = """
    select
        s.emp_id, s.emp_name, s.dept_level_3, s.dept_level_4, s.dept_name, s.level, s.sequence, s.group_code,
        o.diligence_raw_total, o.diligence_month_avg, o.diligence_level,
        o.discipline_raw_count, o.discipline_level,
        o.learning_hours, o.learning_rank_pct, o.learning_level,
        r.final_subjective_grade_1,
        r.final_subjective_grade_2,
        r.final_subjective_grade_3,
        r.current_subjective_level
    from cycle_employee_snapshot s
    left join objective_data o on o.cycle_id = s.cycle_id and o.emp_id = s.emp_id
    left join evaluation_record r on r.cycle_id = s.cycle_id and r.emp_id = s.emp_id
    where s.cycle_id = ? and s.active = 1
    order by s.emp_id
"""


def _none_or_dash(value):
    """数值字段:仅 None 显示为 '-',保留 0 等假值。"""
    return value if value is not None else "-"


def _text_or_dash(value):
    """文本字段:空值(falsy)显示为 '-'。"""
    return value or "-"


def create_process_export(cycle_id: int, operator_id: str, operator_name: str) -> dict:
    """导出过程计算数据:按序列/职级分sheet页，包含客观数据、主观评价和部门负责人确认等级。"""
    rows = get_db().execute(PROCESS_QUERY, (cycle_id,)).fetchall()
    if not rows:
        raise ValueError("no employees to export")

    export_id = f"cycle-{cycle_id}-process"
    export_dir = Path(current_app.config["EXPORT_DIR"]).resolve()
    export_dir.mkdir(parents=True, exist_ok=True)
    file_path = export_dir / f"{export_id}.xlsx"

    workbook = Workbook()
    # 删除默认的sheet
    workbook.remove(workbook.active)

    # 分组数据
    mgmt_rows = []
    p1_p3_rows = []
    p4_p10_rows = []

    for row in rows:
        emp_id = row["emp_id"]
        group_code = row["group_code"]

        # 判断是否属于管理序列（包括强制使用管理维度的人员）
        is_management = (group_code == GROUP_MANAGEMENT) or (emp_id in FORCE_MANAGEMENT_DIMENSIONS_EMPLOYEES)

        if is_management:
            mgmt_rows.append(row)
        elif group_code == GROUP_EMPLOYEE_P1_3:
            p1_p3_rows.append(row)
        elif group_code == GROUP_EMPLOYEE_P4_10:
            p4_p10_rows.append(row)

    # 创建管理序列sheet
    if mgmt_rows:
        _create_sheet(workbook, "管理序列", MGMT_HEADERS, mgmt_rows)

    # 创建P1-P3职级sheet
    if p1_p3_rows:
        _create_sheet(workbook, "P1-P3职级", EMP_HEADERS, p1_p3_rows)

    # 创建P4-P10职级sheet
    if p4_p10_rows:
        _create_sheet(workbook, "P4-P10职级", EMP_HEADERS, p4_p10_rows)

    # 如果所有分组都为空，至少创建一个默认sheet
    if not (mgmt_rows or p1_p3_rows or p4_p10_rows):
        _create_sheet(workbook, "全部数据", EMP_HEADERS, rows)
    workbook.save(file_path)

    write_audit_log(
        action="EXPORT_EXCEL",
        target_type="export_file",
        target_id=export_id,
        operator_id=operator_id,
        operator_name=operator_name,
        cycle_id=cycle_id,
        after_snapshot={"export_id": export_id, "export_type": "process", "file_name": file_path.name},
    )
    get_db().commit()
    return {"export_id": export_id, "file_name": file_path.name, "download_url": f"/exports/{export_id}/download"}


def _create_sheet(workbook: Workbook, title: str, headers: tuple, rows: list) -> None:
    """创建单个sheet并写入数据"""
    sheet = workbook.create_sheet(title=title)
    sheet.append(headers)
    for row in rows:
        sheet.append((
            row["emp_id"],
            row["emp_name"],
            row["dept_level_3"] or row["dept_level_4"] or row["dept_name"] or "-",
            row["dept_level_4"] or row["dept_name"] or "-",
            _text_or_dash(row["level"]),
            _none_or_dash(row["diligence_raw_total"]),
            _none_or_dash(row["diligence_month_avg"]),
            _text_or_dash(row["diligence_level"]),
            _none_or_dash(row["discipline_raw_count"]),
            _text_or_dash(row["discipline_level"]),
            _none_or_dash(row["learning_hours"]),
            _none_or_dash(row["learning_rank_pct"]),
            _text_or_dash(row["learning_level"]),
            _text_or_dash(row["final_subjective_grade_1"]),
            _text_or_dash(row["final_subjective_grade_2"]),
            _text_or_dash(row["final_subjective_grade_3"]),
            _text_or_dash(row["current_subjective_level"]),
        ))


def export_file_path(export_id: str) -> Path:
    return Path(current_app.config["EXPORT_DIR"]).resolve() / f"{export_id}.xlsx"
