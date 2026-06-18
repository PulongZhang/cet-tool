from __future__ import annotations

from io import BytesIO
from typing import BinaryIO

from openpyxl import Workbook, load_workbook

EMPLOYEE_HEADERS = ["工号", "姓名", "一级部门", "二级部门", "三级部门", "四级部门", "岗位", "职级", "序列", "直接上级", "间接上级", "部门负责人"]
OBJECTIVE_HEADERS = ["工号", "勤奋月1", "勤奋月2", "勤奋月3", "考勤异常次数", "日志异常次数", "学习时长"]

EMPLOYEE_FIELD_MAP = {
    "工号": "emp_id",
    "姓名": "emp_name",
    "一级部门": "dept_level_1",
    "二级部门": "dept_level_2",
    "三级部门": "dept_level_3",
    "四级部门": "dept_level_4",
    "岗位": "post",
    "职级": "level",
    "序列": "sequence",
    "直接上级": "direct_manager_id",
    "间接上级": "indirect_manager_id",
    "部门负责人": "dept_head_id",
}
OBJECTIVE_FIELD_MAP = {
    "工号": "emp_id",
    "勤奋月1": "diligence_month_1",
    "勤奋月2": "diligence_month_2",
    "勤奋月3": "diligence_month_3",
    "考勤异常次数": "attendance_exception_count",
    "日志异常次数": "log_exception_count",
    "学习时长": "learning_hours",
}


def build_template(headers: list[str]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "导入模板"
    sheet.append(headers)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def parse_employee_workbook(file: BinaryIO) -> list[dict]:
    return parse_workbook(file, EMPLOYEE_FIELD_MAP)


def parse_objective_workbook(file: BinaryIO) -> list[dict]:
    return parse_workbook(file, OBJECTIVE_FIELD_MAP)


def parse_workbook(file: BinaryIO, field_map: dict[str, str]) -> list[dict]:
    workbook = load_workbook(file, data_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header_indexes = {str(header).strip(): index for index, header in enumerate(header_row) if header}

    missing_headers = [header for header in field_map if header not in header_indexes]
    if missing_headers:
        raise ValueError(f"missing required headers: {', '.join(missing_headers)}")

    rows: list[dict] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if all(value in (None, "") for value in values):
            continue
        row = {}
        for header, field in field_map.items():
            value = values[header_indexes[header]]
            # 将所有值转换为字符串，处理数字类型
            row[field] = str(value) if value is not None else ""
        rows.append(row)
    return rows
