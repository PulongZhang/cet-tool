from performance_app import create_app
from performance_app.db import connect


def make_app(tmp_path):
    return create_app(
        {
            "TESTING": True,
            "DATABASE": str(tmp_path / "app.sqlite3"),
            "EXPORT_DIR": str(tmp_path / "exports"),
        }
    )


def create_cycle(client):
    response = client.post(
        "/cycles",
        json={"cycle_name": "2026-Q2", "start_date": "2026-04-01", "end_date": "2026-06-30"},
    )
    assert response.status_code == 201


def valid_rows():
    return [
        {
            "emp_id": "E001",
            "emp_name": "李四",
            "sequence": "员工序列",
            "level": "P4",
            "dept_name": "平台研发部",
            "direct_manager_id": "M001",
            "indirect_manager_id": "M002",
            "dept_head_id": "M003",
        },
        {
            "emp_id": "M001",
            "emp_name": "张三",
            "sequence": "管理序列",
            "level": "不适用",
            "dept_name": "平台研发部",
            "direct_manager_id": "M002",
            "indirect_manager_id": "M003",
            "dept_head_id": "M003",
        },
        {
            "emp_id": "M002",
            "emp_name": "王经理",
            "sequence": "管理序列",
            "level": "不适用",
            "dept_name": "平台研发部",
            "direct_manager_id": "M003",
            "indirect_manager_id": "M003",
            "dept_head_id": "M003",
        },
        {
            "emp_id": "M003",
            "emp_name": "部门经理",
            "sequence": "管理序列",
            "level": "不适用",
            "dept_name": "平台研发部",
            "direct_manager_id": "",
            "indirect_manager_id": "",
            "dept_head_id": "",
        },
    ]


def test_import_employees_creates_snapshots_records_accounts_and_roles(tmp_path):
    app = make_app(tmp_path)
    client = app.test_client()
    create_cycle(client)

    response = client.post(
        "/cycles/1/employees/import",
        json={"file_name": "employees.xlsx", "rows": valid_rows()},
        headers={"X-Operator-Id": "hr001", "X-Operator-Name": "HR"},
    )

    assert response.status_code == 200
    assert response.get_json()["summary"] == {"total_count": 4, "success_count": 4, "failed_count": 0}

    with connect(app.config["DATABASE"], app.config["DB_ENCRYPTION_KEY"]) as connection:
        snapshots = connection.execute(
            "select emp_id, group_code from cycle_employee_snapshot order by emp_id"
        ).fetchall()
        records = connection.execute(
            "select emp_id, status from evaluation_record order by emp_id"
        ).fetchall()
        role_rows = connection.execute(
            """
            select u.emp_id, r.role_code
            from user_account u
            join user_role r on r.user_id = u.id
            order by u.emp_id, r.role_code
            """
        ).fetchall()

    assert snapshots == [
        ("E001", "EMPLOYEE_P4_10"),
        ("M001", "MANAGEMENT"),
        ("M002", "MANAGEMENT"),
        ("M003", "MANAGEMENT"),
    ]
    assert records == [
        ("E001", "SELF_PENDING"),
        ("M001", "SELF_PENDING"),
        ("M002", "SELF_PENDING"),
        ("M003", "SELF_PENDING"),
    ]
    assert ("E001", "EMPLOYEE") in role_rows
    assert ("M001", "DIRECT_MANAGER") in role_rows
    assert ("M002", "INDIRECT_MANAGER") in role_rows
    assert ("M003", "DEPT_HEAD") in role_rows


def test_import_employees_records_duplicate_and_invalid_level_errors(tmp_path):
    app = make_app(tmp_path)
    client = app.test_client()
    create_cycle(client)
    rows = [
        {
            "emp_id": "E001",
            "emp_name": "李四",
            "sequence": "员工序列",
            "level": "P4",
            "dept_name": "平台研发部",
            "direct_manager_id": "",
            "indirect_manager_id": "",
            "dept_head_id": "",
        },
        {
            "emp_id": "E001",
            "emp_name": "重复员工",
            "sequence": "员工序列",
            "level": "P1",
            "dept_name": "平台研发部",
            "direct_manager_id": "",
            "indirect_manager_id": "",
            "dept_head_id": "",
        },
        {
            "emp_id": "E002",
            "emp_name": "非法职级",
            "sequence": "员工序列",
            "level": "P11",
            "dept_name": "平台研发部",
            "direct_manager_id": "",
            "indirect_manager_id": "",
            "dept_head_id": "",
        },
    ]

    response = client.post("/cycles/1/employees/import", json={"file_name": "employees.xlsx", "rows": rows})

    assert response.status_code == 200
    assert response.get_json()["summary"] == {"total_count": 3, "success_count": 0, "failed_count": 2}
    assert response.get_json()["errors"] == [
        {"row_number": 3, "emp_id": "E001", "field_name": "emp_id", "error_message": "duplicate emp_id in import file"},
        {"row_number": 4, "emp_id": "E002", "field_name": "level", "error_message": "Unsupported employee level: P11"},
    ]

    with connect(app.config["DATABASE"], app.config["DB_ENCRYPTION_KEY"]) as connection:
        errors = connection.execute(
            "select row_number, emp_id, field_name, error_message from import_error order by row_number"
        ).fetchall()
        snapshot_count = connection.execute("select count(*) from cycle_employee_snapshot").fetchone()[0]

    assert errors == [
        (3, "E001", "emp_id", "duplicate emp_id in import file"),
        (4, "E002", "level", "Unsupported employee level: P11"),
    ]
    assert snapshot_count == 0


def test_import_generates_unique_password_per_new_account(tmp_path):
    import csv

    app = make_app(tmp_path)
    client = app.test_client()
    create_cycle(client)

    response = client.post(
        "/cycles/1/employees/import",
        json={"file_name": "employees.xlsx", "rows": valid_rows()},
        headers={"X-Operator-Id": "hr001", "X-Operator-Name": "HR"},
    )
    data = response.get_json()
    assert response.status_code == 200
    assert data["new_account_count"] == 4
    assert data["password_file"]

    download = client.get(data["password_download_url"])
    assert download.status_code == 200
    rows = list(csv.reader(download.data.decode("utf-8-sig").splitlines()))
    assert rows[0] == ["工号", "姓名", "登录账号", "初始密码"]
    body = rows[1:]
    assert len(body) == 4
    passwords = [row[3] for row in body]
    assert len(set(passwords)) == 4  # 每个新员工密码互不相同
    assert {row[0] for row in body} == {"E001", "M001", "M002", "M003"}


def test_reimport_does_not_reset_existing_account_passwords(tmp_path):
    app = make_app(tmp_path)
    client = app.test_client()
    create_cycle(client)
    rows = valid_rows()

    first = client.post(
        "/cycles/1/employees/import",
        json={"file_name": "employees.xlsx", "rows": rows},
        headers={"X-Operator-Id": "hr001"},
    )
    assert first.status_code == 200
    assert first.get_json()["new_account_count"] == 4

    second = client.post(
        "/cycles/1/employees/import",
        json={"file_name": "employees.xlsx", "rows": rows},
        headers={"X-Operator-Id": "hr001"},
    )
    assert second.status_code == 200
    assert second.get_json()["new_account_count"] == 0  # 账号已存在,不再生成新密码


def test_export_cycle_accounts_initial_password_column_points_to_csv(tmp_path):
    from io import BytesIO

    from openpyxl import load_workbook

    app = make_app(tmp_path)
    client = app.test_client()
    create_cycle(client)
    client.post(
        "/cycles/1/employees/import",
        json={"file_name": "employees.xlsx", "rows": valid_rows()},
        headers={"X-Operator-Id": "hr001"},
    )

    response = client.get("/cycles/1/accounts/export")
    assert response.status_code == 200
    sheet = load_workbook(BytesIO(response.data)).active
    header = [cell.value for cell in sheet[1]]
    assert header[-1] == "初始密码"
    # 初始密码不再硬编码 ChangeMe123!,改为提示去查导入批次 CSV
    values = [row[-1].value for row in sheet.iter_rows(min_row=2)]
    assert all("ChangeMe" not in str(v) for v in values)
    assert all("见导入批次密码CSV" in str(v) for v in values)
