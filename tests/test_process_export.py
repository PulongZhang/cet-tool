from io import BytesIO

from openpyxl import load_workbook

from performance_app import create_app
from performance_app.db import connect
from performance_app.services.export_files import create_process_export


def make_app(tmp_path):
    return create_app({
        "TESTING": True,
        "DATABASE": str(tmp_path / "app.sqlite3"),
        "EXPORT_DIR": str(tmp_path / "exports"),
    })


_SNAPSHOT_SQL = """
    insert into cycle_employee_snapshot
        (cycle_id, emp_id, emp_name, sequence, level, group_code, dept_name,
         dept_level_1, dept_level_2, dept_level_3, dept_level_4, post,
         direct_manager_id, indirect_manager_id, dept_head_id, active)
    values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1)
"""


def _create_cycle(app):
    key = app.config["DB_ENCRYPTION_KEY"]
    with connect(app.config["DATABASE"], key) as conn:
        conn.execute(
            "insert into evaluation_cycle (cycle_name, start_date, end_date, status, created_by) "
            "values ('2026-Q2', '2026-04-01', '2026-06-30', 'ACTIVE', 'hr')"
        )
        cycle_id = conn.execute("select id from evaluation_cycle").fetchone()[0]
        conn.commit()
    return cycle_id


def test_export_process_data_contains_objective_raw_values_and_dept_level(tmp_path):
    """P4-P10职级员工导出到'P4-P10职级'sheet页，包含客观数据和主观评价。"""
    app = make_app(tmp_path)
    key = app.config["DB_ENCRYPTION_KEY"]
    cycle_id = _create_cycle(app)

    with connect(app.config["DATABASE"], key) as conn:
        conn.execute(
            _SNAPSHOT_SQL,
            (cycle_id, "E001", "张三", "员工序列", "P4", "EMPLOYEE_P4_10", "研发部",
             "研发部", None, "三级部门", "研发部", "工程师", "", "", ""),
        )
        conn.execute(
            "insert into evaluation_record (cycle_id, emp_id, status, current_subjective_level, "
            "final_subjective_grade_1, final_subjective_grade_2, final_subjective_grade_3) "
            "values (?, 'E001', 'DEPT_CONFIRMED', 'A', 'A', 'B+', 'A')",
            (cycle_id,),
        )
        conn.execute(
            """
            insert into objective_data
                (cycle_id, emp_id, diligence_raw_total, diligence_month_avg, diligence_level,
                 discipline_raw_count, discipline_level, learning_hours, learning_rank_pct, learning_level)
            values (?, 'E001', 180, 60, 'A', 3, 'A+', 12, 100, 'D')
            """,
            (cycle_id,),
        )
        conn.commit()

    with app.app_context():
        result = create_process_export(cycle_id, "hr001", "HR")

    client = app.test_client()
    download = client.get(result["download_url"])
    assert download.status_code == 200
    wb = load_workbook(BytesIO(download.data))

    # 验证sheet页存在
    assert "P4-P10职级" in wb.sheetnames
    sheet = wb["P4-P10职级"]
    rows = [list(r) for r in sheet.iter_rows(values_only=True)]

    assert rows[0] == [
        "工号", "姓名", "三级部门", "四级部门", "职级",
        "勤奋总量", "勤奋月均", "勤奋等级",
        "纪律异常次数", "纪律等级",
        "学习时长(h)", "学习排名%", "学习等级",
        "主观评价-产出质量", "主观评价-主动承担", "主观评价-易用维护",
        "部门负责人确认等级",
    ]
    e001 = [r for r in rows[1:] if r[0] == "E001"][0]
    assert e001[1] == "张三"
    assert e001[2] == "三级部门"  # 三级部门
    assert e001[3] == "研发部"    # 四级部门
    assert e001[4] == "P4"        # 职级
    assert e001[5] == 180         # 勤奋总量(原始值)
    assert e001[6] == 60          # 勤奋月均
    assert e001[7] == "A"         # 勤奋等级
    assert e001[8] == 3           # 纪律异常次数(原始值)
    assert e001[9] == "A+"        # 纪律等级
    assert e001[10] == 12         # 学习时长(原始值)
    assert e001[11] == 100        # 学习排名%
    assert e001[12] == "D"        # 学习等级
    assert e001[13] == "A"        # 主观评价-产出质量
    assert e001[14] == "B+"       # 主观评价-主动承担
    assert e001[15] == "A"        # 主观评价-易用维护
    assert e001[16] == "A"        # 部门负责人确认等级


def test_export_process_data_marks_missing_objective_with_dash(tmp_path):
    """没有客观数据 / 未到部门确认的员工,对应单元格显示 '-'。"""
    app = make_app(tmp_path)
    key = app.config["DB_ENCRYPTION_KEY"]
    cycle_id = _create_cycle(app)

    with connect(app.config["DATABASE"], key) as conn:
        conn.execute(
            _SNAPSHOT_SQL,
            (cycle_id, "E002", "李四", "员工序列", "P5", "EMPLOYEE_P4_10", "研发部",
             "研发部", None, None, "研发部", "", "", "", ""),
        )
        conn.execute(
            "insert into evaluation_record (cycle_id, emp_id, status) values (?, 'E002', 'SELF_PENDING')",
            (cycle_id,),
        )
        conn.commit()

    with app.app_context():
        result = create_process_export(cycle_id, "hr001", "HR")

    client = app.test_client()
    download = client.get(result["download_url"])
    wb = load_workbook(BytesIO(download.data))
    sheet = wb["P4-P10职级"]
    rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    e002 = [r for r in rows[1:] if r[0] == "E002"][0]
    assert e002[5] == "-"    # 勤奋总量(无客观数据)
    assert e002[7] == "-"    # 勤奋等级
    assert e002[13] == "-"   # 主观评价-产出质量(未评价)
    assert e002[14] == "-"   # 主观评价-主动承担(未评价)
    assert e002[15] == "-"   # 主观评价-易用维护(未评价)
    assert e002[16] == "-"   # 部门负责人确认等级(未确认)


def test_export_process_button_downloads_xlsx_directly(tmp_path):
    """点'导出过程计算数据'按钮,POST 直接返回 xlsx 附件,不再需要二次下载链接。"""
    app = make_app(tmp_path)
    key = app.config["DB_ENCRYPTION_KEY"]
    client = app.test_client()
    client.post("/login", data={"username": "hr", "password": "admin123"})

    cycle_id = _create_cycle(app)
    with connect(app.config["DATABASE"], key) as conn:
        conn.execute(
            _SNAPSHOT_SQL,
            (cycle_id, "E001", "张三", "员工序列", "P4", "EMPLOYEE_P4_10", "研发部",
             "研发部", None, "三级部门", "研发部", "工程师", "", "", ""),
        )
        conn.execute(
            "insert into evaluation_record (cycle_id, emp_id, status, current_subjective_level, "
            "final_subjective_grade_1, final_subjective_grade_2, final_subjective_grade_3) "
            "values (?, 'E001', 'DEPT_CONFIRMED', 'A', 'A', 'B+', 'A')",
            (cycle_id,),
        )
        conn.execute(
            """
            insert into objective_data
                (cycle_id, emp_id, diligence_raw_total, diligence_month_avg, diligence_level,
                 discipline_raw_count, discipline_level, learning_hours, learning_rank_pct, learning_level)
            values (?, 'E001', 180, 60, 'A', 3, 'A+', 12, 100, 'D')
            """,
            (cycle_id,),
        )
        conn.commit()

    response = client.post("/page/export-process", data={"cycle_id": str(cycle_id)})

    assert response.status_code == 200
    assert "spreadsheet" in response.content_type
    assert "attachment" in response.headers.get("Content-Disposition", "")
    assert response.data[:2] == b"PK"  # xlsx 是 zip 文件,以 PK 开头


def test_export_management_sequence_has_different_dimension_labels(tmp_path):
    """管理序列导出到'管理序列'sheet页，使用管理维度的标签。"""
    app = make_app(tmp_path)
    key = app.config["DB_ENCRYPTION_KEY"]
    cycle_id = _create_cycle(app)

    with connect(app.config["DATABASE"], key) as conn:
        conn.execute(
            _SNAPSHOT_SQL,
            (cycle_id, "M001", "王经理", "管理序列", "M3", "MANAGEMENT", "管理部",
             "管理部", None, "三级管理部门", "管理部", "经理", "", "", ""),
        )
        conn.execute(
            "insert into evaluation_record (cycle_id, emp_id, status, current_subjective_level, "
            "final_subjective_grade_1, final_subjective_grade_2, final_subjective_grade_3) "
            "values (?, 'M001', 'DEPT_CONFIRMED', 'A', 'A+', 'A', 'B+')",
            (cycle_id,),
        )
        conn.execute(
            """
            insert into objective_data
                (cycle_id, emp_id, diligence_raw_total, diligence_month_avg, diligence_level,
                 discipline_raw_count, discipline_level, learning_hours, learning_rank_pct, learning_level)
            values (?, 'M001', 150, 50, 'A+', 0, 'A+', 20, 95, 'A')
            """,
            (cycle_id,),
        )
        conn.commit()

    with app.app_context():
        result = create_process_export(cycle_id, "hr001", "HR")

    client = app.test_client()
    download = client.get(result["download_url"])
    assert download.status_code == 200
    wb = load_workbook(BytesIO(download.data))

    # 验证sheet页存在
    assert "管理序列" in wb.sheetnames
    sheet = wb["管理序列"]
    rows = [list(r) for r in sheet.iter_rows(values_only=True)]

    # 验证管理序列的主观维度标签
    assert rows[0] == [
        "工号", "姓名", "三级部门", "四级部门", "职级",
        "勤奋总量", "勤奋月均", "勤奋等级",
        "纪律异常次数", "纪律等级",
        "学习时长(h)", "学习排名%", "学习等级",
        "主观评价-工作能力和方法", "主观评价-团队业绩和产出", "主观评价-个人关键任务",
        "部门负责人确认等级",
    ]
    m001 = [r for r in rows[1:] if r[0] == "M001"][0]
    assert m001[1] == "王经理"
    assert m001[2] == "三级管理部门"
    assert m001[3] == "管理部"
    assert m001[4] == "M3"
    assert m001[13] == "A+"  # 主观评价-工作能力和方法
    assert m001[14] == "A"   # 主观评价-团队业绩和产出
    assert m001[15] == "B+"  # 主观评价-个人关键任务
    assert m001[16] == "A"   # 部门负责人确认等级


def test_export_p1_p3_employees_separate_sheet(tmp_path):
    """P1-P3职级员工导出到'P1-P3职级'sheet页。"""
    app = make_app(tmp_path)
    key = app.config["DB_ENCRYPTION_KEY"]
    cycle_id = _create_cycle(app)

    with connect(app.config["DATABASE"], key) as conn:
        # P1职级员工
        conn.execute(
            _SNAPSHOT_SQL,
            (cycle_id, "E101", "小李", "员工序列", "P1", "EMPLOYEE_P1_3", "销售部",
             "销售部", None, "三级销售", "销售部", "专员", "", "", ""),
        )
        conn.execute(
            "insert into evaluation_record (cycle_id, emp_id, status, current_subjective_level, "
            "final_subjective_grade_1, final_subjective_grade_2, final_subjective_grade_3) "
            "values (?, 'E101', 'DEPT_CONFIRMED', 'B+', 'B', 'B+', 'B')",
            (cycle_id,),
        )
        # P4职级员工（应该在P4-P10 sheet）
        conn.execute(
            _SNAPSHOT_SQL,
            (cycle_id, "E401", "小王", "员工序列", "P4", "EMPLOYEE_P4_10", "研发部",
             "研发部", None, "三级研发", "研发部", "工程师", "", "", ""),
        )
        conn.execute(
            "insert into evaluation_record (cycle_id, emp_id, status, current_subjective_level, "
            "final_subjective_grade_1, final_subjective_grade_2, final_subjective_grade_3) "
            "values (?, 'E401', 'DEPT_CONFIRMED', 'A', 'A', 'A', 'B+')",
            (cycle_id,),
        )
        conn.commit()

    with app.app_context():
        result = create_process_export(cycle_id, "hr001", "HR")

    client = app.test_client()
    download = client.get(result["download_url"])
    assert download.status_code == 200
    wb = load_workbook(BytesIO(download.data))

    # 验证sheet页
    assert "P1-P3职级" in wb.sheetnames
    assert "P4-P10职级" in wb.sheetnames

    # 检查P1-P3 sheet
    p1_p3_sheet = wb["P1-P3职级"]
    p1_p3_rows = [list(r) for r in p1_p3_sheet.iter_rows(values_only=True)]
    p1_p3_emp_ids = [r[0] for r in p1_p3_rows[1:]]
    assert "E101" in p1_p3_emp_ids
    assert "E401" not in p1_p3_emp_ids  # P4不在P1-P3 sheet

    # 检查P4-P10 sheet
    p4_p10_sheet = wb["P4-P10职级"]
    p4_p10_rows = [list(r) for r in p4_p10_sheet.iter_rows(values_only=True)]
    p4_p10_emp_ids = [r[0] for r in p4_p10_rows[1:]]
    assert "E401" in p4_p10_emp_ids
    assert "E101" not in p4_p10_emp_ids  # P1不在P4-P10 sheet
