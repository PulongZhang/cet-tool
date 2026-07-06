import sqlite3
from performance_app.db import connect
from io import BytesIO

from openpyxl import Workbook, load_workbook

from performance_app import create_app


def make_app(tmp_path):
    return create_app({"TESTING": True, "DATABASE": str(tmp_path / "app.sqlite3")})


def create_cycle(client):
    client.post("/cycles", json={"cycle_name": "2026-Q2", "start_date": "2026-04-01", "end_date": "2026-06-30"})


def workbook_bytes(headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def employee_rows():
    return [
        ["E001", "李四", "员工序列", "P4", "平台研发部", "M001", "M002", "M003"],
        ["M001", "张三", "管理序列", "不适用", "平台研发部", "M002", "M003", "M003"],
        ["M002", "间接经理", "管理序列", "不适用", "平台研发部", "M003", "M003", "M003"],
        ["M003", "部门负责人", "管理序列", "不适用", "平台研发部", "M003", "M003", "M003"],
    ]


def test_employee_template_and_upload_xlsx(tmp_path):
    app = make_app(tmp_path)
    client = app.test_client()
    create_cycle(client)

    template = client.get("/cycles/1/employees/template")
    assert template.status_code == 200
    workbook = load_workbook(BytesIO(template.data))
    assert list(next(workbook.active.iter_rows(values_only=True))) == [
        "工号",
        "姓名",
        "序列",
        "职级",
        "部门",
        "直接上级工号",
        "间接上级工号",
        "部门负责人工号",
    ]

    upload_file = workbook_bytes(list(next(workbook.active.iter_rows(values_only=True))), employee_rows())
    uploaded = client.post(
        "/cycles/1/employees/upload",
        data={"file": (upload_file, "employees.xlsx")},
        content_type="multipart/form-data",
        headers={"X-Operator-Id": "hr001"},
    )

    assert uploaded.status_code == 200
    assert uploaded.get_json()["summary"] == {"total_count": 4, "success_count": 4, "failed_count": 0}

    with connect(app.config["DATABASE"], app.config["DB_ENCRYPTION_KEY"]) as connection:
        snapshot_count = connection.execute("select count(*) from cycle_employee_snapshot").fetchone()[0]
    assert snapshot_count == 4


def test_objective_template_and_upload_xlsx(tmp_path):
    app = make_app(tmp_path)
    client = app.test_client()
    create_cycle(client)
    client.post(
        "/cycles/1/employees/upload",
        data={
            "file": (
                workbook_bytes(
                    ["工号", "姓名", "序列", "职级", "部门", "直接上级工号", "间接上级工号", "部门负责人工号"],
                    employee_rows(),
                ),
                "employees.xlsx",
            )
        },
        content_type="multipart/form-data",
    )

    template = client.get("/objective/template")
    assert template.status_code == 200
    workbook = load_workbook(BytesIO(template.data))
    assert list(next(workbook.active.iter_rows(values_only=True))) == [
        "工号",
        "勤奋月1",
        "勤奋月2",
        "勤奋月3",
        "考勤异常次数",
        "日志异常次数",
        "学习时长",
    ]

    uploaded = client.post(
        "/objective/upload",
        data={
            "cycle_id": "1",
            "file": (
                workbook_bytes(
                    list(next(workbook.active.iter_rows(values_only=True))),
                    [
                        ["E001", 60, 66, 54, 1, 2, 20],
                        ["M001", 45, 45, 45, 4, 2, 10],
                    ],
                ),
                "objective.xlsx",
            ),
        },
        content_type="multipart/form-data",
        headers={"X-Operator-Id": "hr001", "X-Operator-Name": "HR"},
    )

    assert uploaded.status_code == 200
    assert uploaded.get_json()["summary"] == {"total_count": 2, "success_count": 2, "failed_count": 0}

    with connect(app.config["DATABASE"], app.config["DB_ENCRYPTION_KEY"]) as connection:
        objective_rows = connection.execute(
            "select emp_id, diligence_level, discipline_level from objective_data order by emp_id"
        ).fetchall()
    assert objective_rows == [("E001", "A", "A+"), ("M001", "B", "A")]
